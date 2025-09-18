"""Description: This script manages the excel operations.

Author:
    Julien (@tom4897)
"""

from __future__ import annotations

import logging
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from ..console import ConsoleIO

logger = logging.getLogger(__name__)
ui = ConsoleIO.getUI()  # pylint: disable=invalid-name


@dataclass(slots=True)
class ExcelProcessingMeta:  # pylint: disable=too-many-instance-attributes
    """Generic run metadata to stamp into a workbook."""

    run_at_iso: str
    app_version: str
    source_path: str
    rows_in: int
    rows_out: int
    skipped_rows: int
    errors: int
    warnings: int
    fixes: int
    auto_fix_enabled: bool


class ExcelWorkbookManager:
    """Small utility around openpyxl for reading/writing.

    Small utility around openpyxl for reading/writing:
      - Dataset sheet (header + rows)
      - Key/value config sheet
      - Rules sheet (tabular definitions)
      - Processing metadata + compact report

    Design:
      - Generic (no domain-specific behavior)
      - Explicit lifecycle: call load(), then save()/close()
      - Reads values "as-is" (no coercion beyond header trimming)
    """

    def __init__(self, path: str | Path):
        """Initialize the ExcelWorkbookManager."""
        self.path = Path(path)
        self._wb: Workbook | None = None
        self._opened_from_disk: bool = False

    # Lifecycle
    def load(self, *, data_only: bool = True) -> None:
        """Load workbook. If file doesn't exist, create a new empty workbook. data_only=True reads cached formula values when present."""
        if self.path.exists():
            self._wb = load_workbook(filename=str(self.path), data_only=data_only)
            self._opened_from_disk = True
        else:
            self._wb = Workbook()
            self._opened_from_disk = False

    def save(self, out_path: str | Path | None = None) -> Path:
        """Save workbook to disk. If out_path is None, overwrite original. Returns the path written."""
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        target = Path(out_path) if out_path else self.path
        target.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(target))
        return target

    def close(self) -> None:
        """Release in-memory references."""
        self._wb = None
        self._opened_from_disk = False

    # Reads
    def read_dataset(self, *, sheet: str) -> tuple[list[str], list[list[Any]]]:
        """Returns (header, rows) from `sheet`.

        - Header: first non-empty row (cells coerced to trimmed str)
        - Rows: subsequent non-empty rows (raw values as-is)
        Trailing empties are trimmed to header length.
        """
        ws = self._get_ws(sheet)
        if ws is None:
            ws = self._get_or_create_ws(sheet.lower(), replace=False)
            logger.warning(f"Worksheet '{sheet}' not found in '{self.path.name}'. Creating a new one.")
        if ws is None:
            raise RuntimeError(f"Worksheet '{sheet}' not found in '{self.path.name}'.")
        rows_iter = ws.iter_rows(values_only=True)

        header: list[str] | None = None
        data: list[list[Any]] = []

        # find header
        for raw in rows_iter:
            row = list(raw or [])
            if self._is_empty_row(row):
                continue
            header = [self._normalize_header_cell(c) for c in row]
            break

        if header is None:
            return [], []

        # read rows
        # First pass to count rows for progress tracking
        rows_list = list(rows_iter)
        total_rows = len(rows_list)

        # Add progress tracking if UI is available
        if ui and hasattr(ui, "progress"):
            with ui.progress() as progress:
                task = progress.add_task("Reading Excel data", total=total_rows)

                for raw in rows_list:
                    row = list(raw or [])
                    if self._is_empty_row(row):
                        progress.advance(task)
                        continue
                    if len(row) > len(header):
                        row = row[: len(header)]
                    elif len(row) < len(header):
                        row.extend([None] * (len(header) - len(row)))
                    data.append(row)
                    progress.advance(task)
        else:
            # Fallback without progress tracking
            for raw in rows_list:
                row = list(raw or [])
                if self._is_empty_row(row):
                    continue
                if len(row) > len(header):
                    row = row[: len(header)]
                elif len(row) < len(header):
                    row.extend([None] * (len(header) - len(row)))
                data.append(row)

        return header, data

    def read_config(self, *, sheet: str = "Config") -> dict[str, Any]:
        """Read a two-column key/value sheet.

        A1:'key' B1:'value' header is optional; empty keys skipped.
        """
        ws = self._get_ws(sheet, must_exist=False)
        if ws is None:
            return {}

        cfg: dict[str, Any] = {}
        for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
            if not raw:
                continue
            key = raw[0] if len(raw) > 0 else None
            val = raw[1] if len(raw) > 1 else None

            # ignore header-ish first row
            if i == 1 and str(key).strip().lower() in {"key", "name"}:
                continue

            k = str(key).strip() if key is not None else ""
            if not k:
                continue
            cfg[k] = val
        return cfg

    def read_rules(self, *, sheet: str = "Rules") -> list[dict[str, Any]]:
        """Read a tabular rules sheet into a list of dicts (col_name -> value).

        First non-empty row is header.
        """
        header, rows = self.read_dataset(sheet=sheet)
        if not header:
            return []
        rules: list[dict[str, Any]] = []
        for r in rows:
            if self._is_empty_row(r):
                continue
            rules.append({header[i]: r[i] if i < len(r) else None for i in range(len(header))})
        return rules

    def read_table(self, *, sheet: str, table_name: str) -> list[dict[str, Any]]:
        """Read a structured table from Excel sheet.

        This method supports both Excel's "Format as Table" feature and
        text-based table names for backward compatibility.

        Args:
            sheet: Name of the sheet containing the table
            table_name: Name of the table (e.g., 'CfgAssignees')

        Returns:
            List of dictionaries representing table rows
        """
        ws = self._get_ws(sheet)
        if ws is None:
            logger.warning(f"Sheet '{sheet}' not found")
            return []

        # First, try to find an actual Excel table
        if table_name in ws.tables:
            return self._read_excel_table(ws, table_name)

        # Fallback to text-based table search for backward compatibility
        return self._read_text_based_table(sheet, table_name)

    def _read_excel_table(self, ws, table_name: str) -> list[dict[str, Any]]:
        """Read data from an actual Excel table object.

        Args:
            ws: Worksheet object
            table_name: Name of the Excel table

        Returns:
            List of dictionaries representing table rows
        """
        table = ws.tables[table_name]
        logger.debug(f"Reading Excel table '{table_name}' with ref '{table.ref}'")

        # Get the range of the table
        table_range = ws[table.ref]

        # Extract header row (first row)
        header_row = table_range[0]
        headers = [cell.value for cell in header_row]

        # Extract data rows (remaining rows)
        data_rows = table_range[1:]
        table_data = []

        for row in data_rows:
            row_values = [cell.value for cell in row]
            if self._is_empty_row(row_values):
                continue

            # Create dict from row data using headers
            row_dict = {}
            for i, value in enumerate(row_values):
                if i < len(headers) and headers[i] is not None:
                    row_dict[headers[i]] = value
            table_data.append(row_dict)

        logger.debug(f"Read {len(table_data)} rows from Excel table '{table_name}'")
        return table_data

    def _read_text_based_table(self, sheet: str, table_name: str) -> list[dict[str, Any]]:
        """Read data from text-based table (backward compatibility).

        Args:
            sheet: Name of the sheet containing the table
            table_name: Name of the table to find

        Returns:
            List of dictionaries representing table rows
        """
        header, rows = self.read_dataset(sheet=sheet)
        if not header:
            return []

        # Find the table by looking for the table name in the first column
        table_start = None
        table_columns: list[Any] | None = None

        for i, row in enumerate(rows):
            if row and len(row) > 0 and str(row[0]).strip() == table_name:
                table_start = i
                # The same row contains both table name and column headers
                # Extract only the header columns (skip first cell which holds the table name)
                table_columns = row[1:]
                break

        if table_start is None:
            logger.warning(f"Table '{table_name}' not found in sheet '{sheet}'")
            return []

        if table_columns is None:
            logger.warning(f"Table header row not found for '{table_name}' in sheet '{sheet}'")
            return []

        # Read the table data (skip the table name/header row)
        table_rows = rows[table_start + 1 :]
        table_data = []

        for row in table_rows:
            if self._is_empty_row(row):
                continue
            # Create dict from row data using table header
            row_dict = {}
            # Skip first column (table marker) in data rows and align with columns
            for i, value in enumerate(row[1:]):
                if i < len(table_columns) and table_columns[i] is not None:
                    row_dict[table_columns[i]] = value
            table_data.append(row_dict)

        return table_data

    # Writes

    def write_processing_meta(
        self,
        meta: ExcelProcessingMeta,
        *,
        sheet: str = "_ImportMeta",
        replace: bool = True,
    ) -> None:
        """Write run metadata to a dedicated sheet."""
        ws = self._get_or_create_ws(sheet, replace=replace)
        ws.append(["key", "value"])
        rows = [
            ("run_at_iso", meta.run_at_iso),
            ("app_version", meta.app_version),
            ("source_path", meta.source_path),
            ("rows_in", meta.rows_in),
            ("rows_out", meta.rows_out),
            ("skipped_rows", meta.skipped_rows),
            ("errors", meta.errors),
            ("warnings", meta.warnings),
            ("fixes", meta.fixes),
            ("auto_fix_enabled", meta.auto_fix_enabled),
        ]
        for k, v in rows:
            ws.append([k, v])

    def write_report_table(
        self,
        rows: Iterable[tuple[str, int, str]],
        *,
        sheet: str = "_ImportReport",
        replace: bool = True,
        header: tuple[str, str, str] = ("severity", "count", "code_or_message"),
    ) -> None:
        """Write a compact report table, e.g., aggregated counts by code."""
        ws = self._get_or_create_ws(sheet, replace=replace)
        ws.append(list(header))
        for sev, count, code_or_msg in rows:
            ws.append([sev, count, code_or_msg])

    # Internals
    def _get_ws(self, title: str, *, must_exist: bool = True) -> Worksheet | None:
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        ws = self._wb[title] if title in self._wb.sheetnames else None
        if ws is None and must_exist:
            logger.error(f"Worksheet '{title}' not found in '{self._wb.sheetnames}'.")
            raise KeyError(f"Worksheet '{title}' not found in '{self.path.name}'.")
        return ws

    def _get_or_create_ws(self, title: str, *, replace: bool) -> Worksheet:
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        if title in self._wb.sheetnames:
            if replace:
                index = self._wb.sheetnames.index(title)
                self._wb.remove(self._wb[title])
                return self._wb.create_sheet(title, index=index)
            return self._wb[title]
        return self._wb.create_sheet(title)

    @staticmethod
    def _is_empty_row(row: Iterable[Any]) -> bool:
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, str) and cell.strip() == "":
                continue
            return False
        return True

    @staticmethod
    def _normalize_header_cell(val: Any) -> str:
        # Keep case/spacing under your control; do not lowercase by default.
        return "" if val is None else str(val).strip()
