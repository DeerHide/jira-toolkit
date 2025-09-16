"""
Script Name: excel_io.py
Description: This script manages the excel operations.
Author: Julien (@tom4897)
License: MIT
Date: 2025
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple
import weakref
import logging
from .console import ConsoleIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)
ui = ConsoleIO.getUI()

@dataclass(slots=True)
class ExcelProcessingMeta:
    """
    Generic run metadata to stamp into a workbook.
    Extend as needed; keep it tool-agnostic.
    """
    run_at_iso: str
    app_version: str
    source_path: str
    rows_in: int
    rows_out: int
    skipped_rows: int
    errors: int
    warnings: int
    fixes: int
    auto_fix_enabled: bool


class ExcelWorkbookManager:
    """
    Small utility around openpyxl for reading/writing:
      - Dataset sheet (header + rows)
      - Key/value config sheet
      - Rules sheet (tabular definitions)
      - Processing metadata + compact report

    Design:
      - Generic (no domain-specific behavior)
      - Explicit lifecycle: call load(), then save()/close()
      - Reads values "as-is" (no coercion beyond header trimming)
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self._wb: Optional[Workbook] = None
        self._opened_from_disk: bool = False

    # Lifecycle
    def load(self, *, data_only: bool = True) -> None:
        """
        Load workbook. If file doesn't exist, create a new empty workbook.
        data_only=True reads cached formula values when present.
        """
        if self.path.exists():
            self._wb = load_workbook(filename=str(self.path), data_only=data_only)
            self._opened_from_disk = True
        else:
            self._wb = Workbook()
            self._opened_from_disk = False

    def save(self, out_path: str | Path | None = None) -> Path:
        """
        Save workbook to disk. If out_path is None, overwrite original.
        Returns the path written.
        """
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        target = Path(out_path) if out_path else self.path
        target.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(target))
        return target

    def close(self) -> None:
        """Release in-memory references."""
        self._wb = None
        self._opened_from_disk = False

    # Reads
    def read_dataset(self, *, sheet: str) -> Tuple[List[str], List[List[Any]]]:
        """
        Return (header, rows) from `sheet`.
        - Header: first non-empty row (cells coerced to trimmed str)
        - Rows: subsequent non-empty rows (raw values as-is)
        Trailing empties are trimmed to header length.
        """
        ws = self._get_ws(sheet)
        if ws is None:
            ws = self._get_or_create_ws(sheet.lower())
            logger.warning(f"Worksheet '{sheet}' not found in '{self.path.name}'. Creating a new one.")
        if ws is None:
            raise RuntimeError(f"Worksheet '{sheet}' not found in '{self.path.name}'.")
        rows_iter = ws.iter_rows(values_only=True)

        header: Optional[List[str]] = None
        data: List[List[Any]] = []

        # find header
        for raw in rows_iter:
            row = list(raw or [])
            if self._is_empty_row(row):
                continue
            header = [self._normalize_header_cell(c) for c in row]
            break

        if header is None:
            return [], []

        # read rows
        # First pass to count rows for progress tracking
        rows_list = list(rows_iter)
        total_rows = len(rows_list)

        # Add progress tracking if UI is available
        if ui and hasattr(ui, 'progress'):
            with ui.progress() as progress:
                task = progress.add_task("Reading Excel data", total=total_rows)

                for raw in rows_list:
                    row = list(raw or [])
                    if self._is_empty_row(row):
                        progress.advance(task)
                        continue
                    if len(row) > len(header):
                        row = row[: len(header)]
                    elif len(row) < len(header):
                        row.extend([None] * (len(header) - len(row)))
                    data.append(row)
                    progress.advance(task)
        else:
            # Fallback without progress tracking
            for raw in rows_list:
                row = list(raw or [])
                if self._is_empty_row(row):
                    continue
                if len(row) > len(header):
                    row = row[: len(header)]
                elif len(row) < len(header):
                    row.extend([None] * (len(header) - len(row)))
                data.append(row)

        return header, data

    def read_config(self, *, sheet: str = "Config") -> Dict[str, Any]:
        """
        Read a two-column key/value sheet.
        A1:'key' B1:'value' header is optional; empty keys skipped.
        """
        ws = self._get_ws(sheet, must_exist=False)
        if ws is None:
            return {}

        cfg: Dict[str, Any] = {}
        for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
            if not raw:
                continue
            key = raw[0] if len(raw) > 0 else None
            val = raw[1] if len(raw) > 1 else None

            # ignore header-ish first row
            if i == 1 and str(key).strip().lower() in {"key", "name"}:
                continue

            k = str(key).strip() if key is not None else ""
            if not k:
                continue
            cfg[k] = val
        return cfg

    def read_rules(self, *, sheet: str = "Rules") -> List[Dict[str, Any]]:
        """
        Read a tabular rules sheet into a list of dicts (col_name -> value).
        First non-empty row is header.
        """
        header, rows = self.read_dataset(sheet=sheet)
        if not header:
            return []
        rules: List[Dict[str, Any]] = []
        for r in rows:
            if self._is_empty_row(r):
                continue
            rules.append({header[i]: r[i] if i < len(r) else None for i in range(len(header))})
        return rules

    # Writes

    def write_processing_meta(
        self,
        meta: ExcelProcessingMeta,
        *,
        sheet: str = "_ImportMeta",
        replace: bool = True,
    ) -> None:
        """
        Write run metadata to a dedicated sheet.
        """
        ws = self._get_or_create_ws(sheet, replace=replace)
        ws.append(["key", "value"])
        rows = [
            ("run_at_iso", meta.run_at_iso),
            ("app_version", meta.app_version),
            ("source_path", meta.source_path),
            ("rows_in", meta.rows_in),
            ("rows_out", meta.rows_out),
            ("skipped_rows", meta.skipped_rows),
            ("errors", meta.errors),
            ("warnings", meta.warnings),
            ("fixes", meta.fixes),
            ("auto_fix_enabled", meta.auto_fix_enabled),
        ]
        for k, v in rows:
            ws.append([k, v])

    def write_report_table(
        self,
        rows: Iterable[Tuple[str, int, str]],
        *,
        sheet: str = "_ImportReport",
        replace: bool = True,
        header: Tuple[str, str, str] = ("severity", "count", "code_or_message"),
    ) -> None:
        """
        Write a compact report table, e.g., aggregated counts by code.
        """
        ws = self._get_or_create_ws(sheet, replace=replace)
        ws.append(list(header))
        for sev, count, code_or_msg in rows:
            ws.append([sev, count, code_or_msg])

    # Internals

    def _get_ws(self, title: str, *, must_exist: bool = True) -> Optional[Worksheet]:
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        ws = self._wb[title] if title in self._wb.sheetnames else None
        if ws is None and must_exist:
            logger.error(f"Worksheet '{title}' not found in '{self._wb.sheetnames}'.")
            raise KeyError(f"Worksheet '{title}' not found in '{self.path.name}'.")
        return ws

    def _get_or_create_ws(self, title: str, *, replace: bool) -> Worksheet:
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        if title in self._wb.sheetnames:
            if replace:
                index = self._wb.sheetnames.index(title)
                self._wb.remove(self._wb[title])
                return self._wb.create_sheet(title, index=index)
            return self._wb[title]
        return self._wb.create_sheet(title)

    @staticmethod
    def _is_empty_row(row: Iterable[Any]) -> bool:
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, str) and cell.strip() == "":
                continue
            return False
        return True

    @staticmethod
    def _normalize_header_cell(val: Any) -> str:
        # Keep case/spacing under your control; do not lowercase by default.
        return "" if val is None else str(val).strip()
