"""Tests for workbook-level config table discovery and indexed lookup."""

from __future__ import annotations

from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from jira_importer.config.excel_config import ExcelConfiguration  # type: ignore[import-untyped]
from jira_importer.errors import ConfigurationError, ErrorCode  # type: ignore[import-untyped]
from jira_importer.excel.excel_io import ExcelWorkbookManager  # type: ignore[import-untyped]
from jira_importer.excel.excel_table_reader import ExcelTableReader  # type: ignore[import-untyped]


def _active_ws(wb: Workbook) -> Worksheet:
    ws = wb.active
    return cast("Worksheet", ws)


def _write_table(ws: Worksheet, table_name: str, headers: list[str], row_values: list[Any], start_row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
    for col_idx, value in enumerate(row_values, start=1):
        ws.cell(row=start_row + 1, column=col_idx, value=value)

    end_col_letter = chr(ord("A") + len(headers) - 1)
    table_ref = f"A{start_row}:{end_col_letter}{start_row + 1}"
    table = Table(displayName=table_name, ref=table_ref)
    ws.add_table(table)


def _save_workbook(path: Path, setup: Any) -> None:
    wb = Workbook()
    setup(wb)
    wb.save(path)


def test_indexed_read_discovers_tables_across_case_insensitive_config_prefixes(tmp_path: Path) -> None:
    """Discover tables only from case-insensitive config-prefixed sheets."""
    file_path = tmp_path / "multi_config.xlsx"

    def setup(wb: Workbook) -> None:
        ws_primary = _active_ws(wb)
        ws_primary.title = "Config_Main"
        _write_table(
            ws_primary,
            "CfgAssignees",
            ["Assignee.Name", "Assignee.ID"],
            ["Jane Doe", "acct-1"],
        )

        ws_secondary = wb.create_sheet("CFG-Team")
        _write_table(
            ws_secondary,
            "CfgSprints",
            ["Sprint.Name", "Sprint.ID"],
            ["Sprint 1", "101"],
        )

        ws_ignored = wb.create_sheet("DataSheet")
        _write_table(
            ws_ignored,
            "CfgFixVersions",
            ["FixVersion.Name"],
            ["1.0"],
        )

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        assignees = manager.read_table(sheet=None, table_name="CfgAssignees")
        sprints = manager.read_table(sheet=None, table_name="CfgSprints")

        assert assignees == [{"Assignee.Name": "Jane Doe", "Assignee.ID": "acct-1"}]
        assert sprints == [{"Sprint.Name": "Sprint 1", "Sprint.ID": "101"}]

        with pytest.raises(ConfigurationError):
            manager.read_table(sheet=None, table_name="CfgFixVersions")
    finally:
        manager.close()


def test_read_table_explicit_sheet_mode_remains_supported(tmp_path: Path) -> None:
    """Keep explicit sheet lookups backward-compatible."""
    file_path = tmp_path / "single_config.xlsx"

    def setup(wb: Workbook) -> None:
        ws = _active_ws(wb)
        ws.title = "Config"
        _write_table(
            ws,
            "CfgIssueTypes",
            ["IssueType.Name"],
            ["Story"],
        )

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        issue_types = manager.read_table(sheet="Config", table_name="CfgIssueTypes")
        assert issue_types == [{"IssueType.Name": "Story"}]
    finally:
        manager.close()


def test_missing_required_table_error_contains_structured_details(tmp_path: Path) -> None:
    """Raise structured error details for missing required indexed tables."""
    file_path = tmp_path / "missing_required.xlsx"

    def setup(wb: Workbook) -> None:
        ws = _active_ws(wb)
        ws.title = "config-main"

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        with pytest.raises(ConfigurationError) as exc_info:
            manager.read_table(sheet=None, table_name="CfgPriorities")

        details = exc_info.value.details
        assert details["table_name"] == "CfgPriorities"
        assert details["mode"] == "indexed"
        assert details["candidate_sheets"] == ["config-main"]
        assert details["n_candidate_sheets"] == 1
        assert str(file_path) == details["workbook_path"]
    finally:
        manager.close()


def test_config_table_index_resets_on_close_and_reloads_once(tmp_path: Path) -> None:
    """Build index lazily and reset it on close."""
    file_path = tmp_path / "index_lifecycle.xlsx"

    def setup(wb: Workbook) -> None:
        ws = _active_ws(wb)
        ws.title = "Config"
        _write_table(
            ws,
            "CfgIgnoreList",
            ["IgnoreList.Name"],
            ["skip"],
        )

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        assert not manager.is_config_table_index_built()
        manager.read_table(sheet=None, table_name="CfgIgnoreList")
        assert manager.is_config_table_index_built()
        manager.read_table(sheet=None, table_name="CfgIgnoreList")
        assert manager.is_config_table_index_built()
    finally:
        manager.close()

    assert not manager.is_config_table_index_built()


def test_excel_table_reader_uses_indexed_discovery_for_default_config(tmp_path: Path) -> None:
    """Read all required config tables across config-prefixed sheets."""
    file_path = tmp_path / "reader_indexed.xlsx"

    def setup(wb: Workbook) -> None:
        ws1 = _active_ws(wb)
        ws1.title = "Config_Main"
        _write_table(ws1, "CfgAssignees", ["Assignee.Name", "Assignee.ID"], ["Ann", "acc-10"])
        _write_table(ws1, "CfgSprints", ["Sprint.Name", "Sprint.ID"], ["S1", "11"], start_row=4)
        _write_table(ws1, "CfgFixVersions", ["FixVersion.Name"], ["1.0"], start_row=7)
        _write_table(ws1, "CfgComponents", ["Component.Name"], ["Backend"], start_row=10)

        ws2 = wb.create_sheet("cfg_02")
        _write_table(ws2, "CfgIssueTypes", ["IssueType.Name"], ["Task"])
        _write_table(ws2, "CfgIgnoreList", ["IgnoreList.Name"], ["note"], start_row=4)
        _write_table(ws2, "CfgPriorities", ["Priority.Name"], ["High"], start_row=7)
        _write_table(ws2, "CfgAutofieldValues", ["Name", "Value"], ["jira.connection.timeout", "30"], start_row=10)

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        reader = ExcelTableReader(manager)
        tables = reader.read_all_tables(config_sheet="Config")
        assert tables.assignees[0].name == "Ann"
        assert tables.sprints[0].id == "11"
        assert tables.priorities[0].name == "High"
        assert tables.auto_field_values[0].name == "jira.connection.timeout"
    finally:
        manager.close()


def test_indexed_text_table_found_with_no_rows_is_valid_empty(tmp_path: Path) -> None:
    """Treat found text-based table with zero rows as valid empty table."""
    file_path = tmp_path / "empty_text_table.xlsx"

    def setup(wb: Workbook) -> None:
        ws = _active_ws(wb)
        ws.title = "cfg-main"
        # Dataset header row used by read_dataset()
        ws.cell(row=1, column=1, value="Section")
        ws.cell(row=1, column=2, value="Col1")
        # Text-based table marker/header row (no data rows after it)
        ws.cell(row=2, column=1, value="CfgIssueTypes")
        ws.cell(row=2, column=2, value="IssueType.Name")

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        issue_types = manager.read_table(sheet=None, table_name="CfgIssueTypes", optional=False)
        assert issue_types == []
    finally:
        manager.close()


def test_excel_configuration_cached_table_load_failure_stops_retries(tmp_path: Path) -> None:
    """Avoid reloading table config after a prior hard failure."""
    file_path = tmp_path / "no_issue_types.xlsx"

    def setup(wb: Workbook) -> None:
        ws_cfg = _active_ws(wb)
        ws_cfg.title = "Config"
        ws_cfg.cell(row=1, column=1, value="key")
        ws_cfg.cell(row=1, column=2, value="value")
        ws_cfg.cell(row=2, column=1, value="metadata.version")
        ws_cfg.cell(row=2, column=2, value=7)

        ws_table = wb.create_sheet("cfg-main")
        _write_table(
            ws_table,
            "CfgAssignees",
            ["Assignee.Name", "Assignee.ID"],
            ["User A", "acc-1"],
        )
        # Intentionally omit CfgIssueTypes to trigger required-table failure.

    _save_workbook(file_path, setup)

    # Initialization now fails fast when required config tables are missing.
    with pytest.raises(Exception) as exc_info:
        ExcelConfiguration(str(file_path), config_sheet="Config")
    assert "Required table" in str(exc_info.value)
    assert isinstance(exc_info.value, ConfigurationError)
    assert exc_info.value.code == ErrorCode.CONFIG_MISSING_REQUIRED


def test_excel_table_reader_treats_sprint_fixversion_component_tables_as_optional(tmp_path: Path) -> None:
    """Allow missing CfgSprints/CfgFixVersions/CfgComponents without failing table config load."""
    file_path = tmp_path / "optional_tables_missing.xlsx"

    def setup(wb: Workbook) -> None:
        ws_cfg = _active_ws(wb)
        ws_cfg.title = "cfg-main"
        _write_table(ws_cfg, "CfgAssignees", ["Assignee.Name", "Assignee.ID"], ["Ann", "acc-10"])
        _write_table(ws_cfg, "CfgIssueTypes", ["IssueType.Name"], ["Story"], start_row=4)
        _write_table(ws_cfg, "CfgIgnoreList", ["IgnoreList.Name"], ["note"], start_row=7)
        _write_table(ws_cfg, "CfgPriorities", ["Priority.Name"], ["High"], start_row=10)
        _write_table(ws_cfg, "CfgAutofieldValues", ["Name", "Value"], ["jira.connection.timeout", "30"], start_row=13)
        # Intentionally omit CfgSprints, CfgFixVersions, and CfgComponents.

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        reader = ExcelTableReader(manager)
        tables = reader.read_all_tables(config_sheet="Config")
        assert tables.assignees[0].name == "Ann"
        assert tables.sprints == []
        assert tables.fix_versions == []
        assert tables.components == []
    finally:
        manager.close()


def test_excel_table_reader_reads_cfgsettings_table(tmp_path: Path) -> None:
    """Read CfgSettings rows when present in config-prefixed sheets."""
    file_path = tmp_path / "settings_table.xlsx"

    def setup(wb: Workbook) -> None:
        ws_cfg = _active_ws(wb)
        ws_cfg.title = "cfg-main"
        _write_table(ws_cfg, "CfgAssignees", ["Assignee.Name", "Assignee.ID"], ["Ann", "acc-10"])
        _write_table(ws_cfg, "CfgIssueTypes", ["IssueType.Name"], ["Story"], start_row=4)
        _write_table(ws_cfg, "CfgIgnoreList", ["IgnoreList.Name"], ["note"], start_row=7)
        _write_table(ws_cfg, "CfgPriorities", ["Priority.Name"], ["High"], start_row=10)
        _write_table(ws_cfg, "CfgAutofieldValues", ["Name", "Value"], ["jira.connection.timeout", "30"], start_row=13)
        _write_table(
            ws_cfg,
            "CfgSettings",
            ["Name", "Value", "Type"],
            ["metadata.version", "7", "int"],
            start_row=16,
        )

    _save_workbook(file_path, setup)

    manager = ExcelWorkbookManager(file_path)
    manager.load()
    try:
        reader = ExcelTableReader(manager)
        tables = reader.read_all_tables(config_sheet="Config")
        assert len(tables.settings) == 1
        assert tables.settings[0].name == "metadata.version"
        assert tables.settings[0].value == 7
        assert tables.settings[0].value_type == "int"
    finally:
        manager.close()
